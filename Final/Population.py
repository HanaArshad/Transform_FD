from pathlib import Path
from openpyxl.styles import PatternFill

input_dir = Path.cwd() / '/Users/hanaarshadahmed/Desktop/NEW'
from openpyxl import load_workbook  # pip install openpyxl
for path in list(input_dir.rglob("*.xlsx*")):
    wb = load_workbook(filename=path)
    ws = wb['Contents']
    del wb['Contents']
    ws1 = wb['Dataset Info']
    del wb['Dataset Info']
    ws2 = wb['INC']
    del wb['INC']
    ws3 = wb['ECON']
    del wb['ECON']
    ws4 = wb['EDU']
    del wb['EDU']
    ws5 = wb['HEAL']
    del wb['HEAL']
    ws6 = wb['FAM']
    del wb['FAM']
    ws7 = wb['MIG']
    del wb['MIG']
    ws8 = wb['ENV']
    del wb['ENV']
    ws9 = wb['ING_POP']
    del wb['ING_POP']

    sheet = wb['POP']
    rows_to_delete = [1, 2, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27,
                      28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51,
                      52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75,
                      76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99,
                      100, 101, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117, 118,119,
                      120, 121, 122, 123, 124, 127, 128, 129, 130, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140,
                      141,142, 143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 156, 157]
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_index)
    sheet.delete_cols(1,2)
    # Remove colour in cell
    for rows in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2000):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(fill_type="none")

    output_dir = Path.cwd() / '/Users/hanaarshadahmed/Desktop/POP_Final'
    output_dir.mkdir(exist_ok=True)
    wb.save(output_dir / path.name)