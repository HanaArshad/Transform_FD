import pandas as pd
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill

input_dir = Path.cwd() / '/Users/hanaarshadahmed/Desktop/NEW'
from openpyxl import load_workbook  # pip install openpyxl

for path in list(input_dir.rglob("*.xlsx*")):
    wb_obj = load_workbook(filename=path)
    # Give the location of the file
    wb_obj = openpyxl.load_workbook(path)
    del wb_obj['Contents']
    del wb_obj['Dataset Info']
    del wb_obj['POP']
    del wb_obj['ING_POP']
    del wb_obj['ECON']
    del wb_obj['EDU']
    del wb_obj['HEAL']
    del wb_obj['FAM']
    del wb_obj['MIG']
    del wb_obj['ENV']

    sheet = wb_obj['INC']
    rows_to_delete = [1,2,4,5,6,9,10,11,12,15,16,17,18,21,22,23,24,27,28,29,30,33,34,35,36,37,38,39,40,41,42,43,
                      44,45,46,47,48,49,50,51,52,53,54,55,56, 57,58,59,60,61,62,63,64,65,66,67,68,70,71,72,73,74,
                      75,76,77]
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_index)
    # Remove colour in cell
    for rows in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2000):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(fill_type="none")


    sheet.delete_cols(1,2)

    output_dir = Path.cwd() / '/Users/hanaarshadahmed/Desktop/Personal_Income'
    output_dir.mkdir(exist_ok=True)
    wb_obj.save(output_dir / path.name)